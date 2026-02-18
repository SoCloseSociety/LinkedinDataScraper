"""Export LinkedIn profiles to CSV and Excel with professional formatting."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from linkedin_scraper.models import LinkedInProfile

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions grouped by category
# ---------------------------------------------------------------------------
COLUMNS = [
    # (internal_key, display_label, category)
    ("full_name", "Full Name", "identity"),
    ("headline", "Headline", "identity"),
    ("current_company", "Company", "identity"),
    ("location", "Location", "identity"),
    ("industry", "Industry", "identity"),
    ("email", "Email", "contact"),
    ("phone", "Phone", "contact"),
    ("website", "Website", "contact"),
    ("profile_url", "LinkedIn URL", "contact"),
    ("current_title", "Current Title", "professional"),
    ("experience_summary", "Experience", "professional"),
    ("education_summary", "Education", "professional"),
    ("skills_summary", "Top Skills", "professional"),
    ("connections_count", "Connections", "professional"),
    ("about", "About", "about"),
    ("data_source", "Source", "meta"),
    ("search_query", "Search Query", "meta"),
    ("search_location", "Search Location", "meta"),
    ("scraped_at", "Scraped At", "meta"),
]

COLUMN_KEYS = [c[0] for c in COLUMNS]
COLUMN_LABELS = {c[0]: c[1] for c in COLUMNS}
COLUMN_CATEGORIES = {c[0]: c[2] for c in COLUMNS}

# ---------------------------------------------------------------------------
# Excel color scheme — SoClose brand colors
# Primary: #575ECF (purple), Dark: #1b1b1b, Light: #7B80E0
# ---------------------------------------------------------------------------
_HEADER_FILLS = {
    "identity": PatternFill(start_color="575ECF", end_color="575ECF", fill_type="solid"),   # SoClose purple
    "contact": PatternFill(start_color="7B80E0", end_color="7B80E0", fill_type="solid"),     # SoClose light purple
    "professional": PatternFill(start_color="1B1B1B", end_color="1B1B1B", fill_type="solid"),# SoClose dark
    "about": PatternFill(start_color="3D3D3D", end_color="3D3D3D", fill_type="solid"),       # Dark gray
    "meta": PatternFill(start_color="86888A", end_color="86888A", fill_type="solid"),         # Neutral gray
}

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
FONT_DEFAULT = Font(size=10, name="Calibri")
FONT_LINK = Font(color="575ECF", underline="single", size=10, name="Calibri")  # SoClose purple links
FONT_BOLD = Font(bold=True, size=10, name="Calibri")
FONT_NAME = Font(bold=True, size=11, name="Calibri", color="1B1B1B")  # SoClose dark
FONT_DIM = Font(size=9, name="Calibri", color="666666")

FILL_HAS_EMAIL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NO_EMAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")     # SoClose soft bg
FILL_HAS_DATA = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_HAS_PHONE = PatternFill(start_color="E8E9F7", end_color="E8E9F7", fill_type="solid")   # Light purple tint

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALIGNMENT_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")

# Recommended column widths
_COL_WIDTHS = {
    "full_name": 24,
    "headline": 38,
    "current_company": 26,
    "location": 22,
    "industry": 22,
    "email": 32,
    "phone": 20,
    "website": 32,
    "profile_url": 42,
    "current_title": 28,
    "experience_summary": 55,
    "education_summary": 42,
    "skills_summary": 38,
    "connections_count": 14,
    "about": 55,
    "data_source": 10,
    "search_query": 20,
    "search_location": 18,
    "scraped_at": 22,
}


# ======================================================================
# Public API
# ======================================================================

def export_profiles(
    profiles: list[LinkedInProfile],
    output_dir: str = "output",
    fmt: str = "both",
    keywords: str = "",
    location: str = "",
) -> list[Path]:
    """Export profiles to CSV and/or Excel. Returns paths of created files."""
    if not profiles:
        logger.warning("No profiles to export.")
        return []

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _safe_filename(f"linkedin_{keywords}_{location}_{timestamp}")

    df = _prepare_dataframe(profiles)
    created: list[Path] = []

    if fmt in ("csv", "both"):
        csv_path = out / f"{base}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        created.append(csv_path)
        logger.info("CSV exported: %s (%d rows)", csv_path, len(df))

    if fmt in ("excel", "both"):
        xlsx_path = out / f"{base}.xlsx"
        _export_excel(df, xlsx_path, keywords, location)
        created.append(xlsx_path)
        logger.info("Excel exported: %s (%d rows)", xlsx_path, len(df))

    return created


# ======================================================================
# DataFrame preparation
# ======================================================================

def _prepare_dataframe(profiles: list[LinkedInProfile]) -> pd.DataFrame:
    """Flatten profiles into a tabular DataFrame."""
    rows = []
    for p in profiles:
        row = {
            "full_name": p.full_name or "",
            "headline": p.headline or "",
            "current_company": p.current_company or "",
            "location": p.location or "",
            "industry": p.industry or "",
            "email": p.email or "",
            "phone": p.phone or "",
            "website": p.website or "",
            "profile_url": p.profile_url or "",
            "current_title": p.current_title or "",
            "experience_summary": p.experience_summary(),
            "education_summary": p.education_summary(),
            "skills_summary": p.skills_summary(),
            "connections_count": p.connections_count or "",
            "about": (p.about or "")[:500],
            "data_source": p.data_source or "",
            "search_query": p.search_query or "",
            "search_location": p.search_location or "",
            "scraped_at": p.scraped_at or "",
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMN_KEYS)

    # Replace NaN and None with empty strings to avoid "None" in Excel
    df = df.fillna("")

    # Rename columns to display labels
    df.rename(columns=COLUMN_LABELS, inplace=True)

    # Sort: profiles with email first, then by name
    email_col = COLUMN_LABELS["email"]
    name_col = COLUMN_LABELS["full_name"]
    df["_has_email"] = df[email_col].apply(lambda x: 0 if x and str(x).strip() else 1)
    df.sort_values(["_has_email", name_col], inplace=True)
    df.drop(columns=["_has_email"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


# ======================================================================
# Excel export with formatting
# ======================================================================

def _export_excel(df: pd.DataFrame, path: Path, keywords: str = "", location: str = "") -> None:
    """Write a professionally formatted Excel file."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="LinkedIn Profiles")
        ws = writer.sheets["LinkedIn Profiles"]

        num_rows = len(df) + 1  # +1 for header
        num_cols = len(df.columns)

        # --- Row height ---
        ws.row_dimensions[1].height = 30  # Taller header row

        # --- Header formatting ---
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            key = COLUMN_KEYS[col_idx - 1]
            category = COLUMN_CATEGORIES.get(key, "meta")

            cell.font = FONT_HEADER
            cell.fill = _HEADER_FILLS.get(category, _HEADER_FILLS["meta"])
            cell.alignment = ALIGNMENT_CENTER
            cell.border = THIN_BORDER

        # --- Column widths ---
        for col_idx in range(1, num_cols + 1):
            key = COLUMN_KEYS[col_idx - 1]
            width = _COL_WIDTHS.get(key, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # --- Precompute column indices ---
        name_col_idx = COLUMN_KEYS.index("full_name") + 1
        email_col_idx = COLUMN_KEYS.index("email") + 1
        phone_col_idx = COLUMN_KEYS.index("phone") + 1
        url_col_idx = COLUMN_KEYS.index("profile_url") + 1
        website_col_idx = COLUMN_KEYS.index("website") + 1
        source_col_idx = COLUMN_KEYS.index("data_source") + 1

        # --- Data rows ---
        for row_idx in range(2, num_rows + 1):
            is_alt_row = (row_idx - 2) % 2 == 1

            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = FONT_DEFAULT
                cell.border = THIN_BORDER
                cell.alignment = ALIGNMENT_WRAP

                # Alternating row background
                if is_alt_row:
                    cell.fill = FILL_ALT_ROW

            # --- Name column: bold ---
            name_cell = ws.cell(row=row_idx, column=name_col_idx)
            name_cell.font = FONT_NAME

            # --- Email cell color coding ---
            email_cell = ws.cell(row=row_idx, column=email_col_idx)
            email_val = str(email_cell.value or "").strip()
            if email_val and email_val != "None":
                email_cell.fill = FILL_HAS_EMAIL
                email_cell.font = FONT_BOLD
            else:
                email_cell.fill = FILL_NO_EMAIL
                email_cell.value = ""  # Clear "None" values

            # --- Phone cell color coding ---
            phone_cell = ws.cell(row=row_idx, column=phone_col_idx)
            phone_val = str(phone_cell.value or "").strip()
            if phone_val and phone_val != "None":
                phone_cell.fill = FILL_HAS_PHONE
            else:
                phone_cell.value = ""

            # --- Website cell: clear None ---
            website_cell = ws.cell(row=row_idx, column=website_col_idx)
            website_val = str(website_cell.value or "").strip()
            if website_val == "None":
                website_cell.value = ""
                website_val = ""
            if website_val.startswith("http"):
                website_cell.hyperlink = website_val
                website_cell.font = FONT_LINK

            # --- Clickable LinkedIn URL ---
            url_cell = ws.cell(row=row_idx, column=url_col_idx)
            url_val = str(url_cell.value or "").strip()
            if url_val.startswith("http"):
                url_cell.hyperlink = url_val
                url_cell.font = FONT_LINK

            # --- Source column: center aligned, dim font ---
            source_cell = ws.cell(row=row_idx, column=source_col_idx)
            source_cell.alignment = ALIGNMENT_CENTER
            source_cell.font = FONT_DIM

            # --- Clean any remaining "None" values in all cells ---
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "None":
                    cell.value = ""

        # --- Freeze panes (header row + first column) ---
        ws.freeze_panes = "B2"

        # --- Auto filter ---
        ws.auto_filter.ref = ws.dimensions

        # --- Summary sheet ---
        _add_summary_sheet(writer, df, keywords, location)

    logger.info("Excel file written: %s (%d profiles)", path, len(df))


def _add_summary_sheet(writer, df: pd.DataFrame, keywords: str, location: str) -> None:
    """Add a summary/statistics sheet to the Excel file."""
    wb = writer.book
    ws = wb.create_sheet("Summary", 0)  # Insert at beginning

    # Title — SoClose branded
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="LinkedIn Data Scraper — Report")
    title_cell.font = Font(bold=True, size=16, color="575ECF", name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle — SoClose branding
    ws.merge_cells("A2:D2")
    sub_cell = ws.cell(row=2, column=1, value="Generated by SoClose | soclose.co")
    sub_cell.font = Font(size=10, color="999999", name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center")

    # Metadata
    row = 4
    email_col = COLUMN_LABELS["email"]
    phone_col = COLUMN_LABELS["phone"]
    source_col = COLUMN_LABELS["data_source"]

    total_profiles = len(df)
    with_email = int((df[email_col].astype(str).str.strip() != "").sum())
    with_phone = int((df[phone_col].astype(str).str.strip() != "").sum())
    from_api = int((df[source_col] == "api").sum())
    from_profile = int((df[source_col] == "profile").sum())
    from_search = int((df[source_col] == "search").sum())

    stats = [
        ("Search Keywords", keywords or "N/A"),
        ("Location", location or "Any"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Profiles", total_profiles),
        ("With Email", f"{with_email} ({_pct(with_email, total_profiles)})"),
        ("With Phone", f"{with_phone} ({_pct(with_phone, total_profiles)})"),
        ("", ""),
        ("Data from API", f"{from_api} ({_pct(from_api, total_profiles)})"),
        ("Data from Profile page", f"{from_profile} ({_pct(from_profile, total_profiles)})"),
        ("Data from Search only", f"{from_search} ({_pct(from_search, total_profiles)})"),
    ]

    label_font = Font(bold=True, size=11, name="Calibri", color="333333")
    value_font = Font(size=11, name="Calibri")

    for label, value in stats:
        if label:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value).font = value_font
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35

    # Set the LinkedIn Profiles sheet as the active/default sheet
    wb.active = wb.sheetnames.index("LinkedIn Profiles")


def _pct(part: int, total: int) -> str:
    """Format as percentage string."""
    if total == 0:
        return "0%"
    return f"{part * 100 // total}%"


# ======================================================================
# Helpers
# ======================================================================

def _safe_filename(text: str) -> str:
    """Sanitize text for use as a filename."""
    safe = text.strip().lower()
    safe = re.sub(r"[^\w\s-]", "", safe)
    safe = re.sub(r"[\s]+", "_", safe)
    return safe[:100]
